from itertools import product
from openpyxl import Workbook

price_value_pairs = [
    (124, 500, 3),
    (249, 1050, 3),
    (495, 2175, 3),
    (870, 3850, 3),
    (1235, 5550, 3),
    (2480, 11500, 10),
]


def calculate_combinations(pairs):
    results = {}
    details = []

    counts_ranges = [range(max_count + 1) for _, _, max_count in pairs]

    for counts in product(*counts_ranges):
        if 0 < sum(counts):
            total_price = sum(p * c for (p, _, _), c in zip(pairs, counts))
            total_value = sum(v * c for (_, v, _), c in zip(pairs, counts))
            if total_price not in results or total_value > results[total_price][1]:
                results[total_price] = (counts, total_value)

    for price, (counts, value) in results.items():
        details.append((price, value, counts))

    return details


def create_excel_file(details, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Valorant"

    ws.cell(row=1, column=1, value="ราคาบาท")
    ws.cell(row=1, column=2, value="จำนวนที่ได้")
    ws.cell(row=1, column=3, value="แพ็คผสม(บาท)")

    for idx, (price, value, counts) in enumerate(details, start=2):
        count_detail = ' + '.join(
            [f"{pairs[0]}" for pairs, count in zip(price_value_pairs, counts) if count > 0 for _ in range(count)])
        ws.cell(row=idx, column=1, value=price)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=3, value=count_detail)

    wb.save(filename)
    print(f"บันทึกไฟล์ Excel เรียบร้อยที่ {filename}")


details = calculate_combinations(price_value_pairs)
print(f"มีทั้งหมด: {len(details)} รูปแบบ")
print("""นี่คือราคาทั้งหมด
🔽 🔽 🔽 🔽""")

sorted_details = sorted(details, key=lambda x: (x[0], x[1]))

for price, value, counts in sorted_details:
    count_detail = ' + '.join(
        [f"{pairs[0]}" for pairs, count in zip(price_value_pairs, counts) if count > 0 for _ in range(count)])
    print(f"ราคา {price} บาท ได้ {value} แพ็ค: {count_detail}")

save_path = r'C:\Users\Core i5\Desktop\Price\Valorant_TUN.xlsx'
create_excel_file(sorted_details, save_path)
